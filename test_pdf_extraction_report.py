import os
import pandas as pd
import re
from datetime import datetime
from pdf_header_extractor_v2 import extract_header_content
from header_data_filter import filter_header_data
import openpyxl

# ANSI color codes
class Colors:
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BLUE = '\033[94m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'

def print_colored(text, color):
    """Print text with color"""
    print(f"{color}{text}{Colors.ENDC}")

def print_section(title):
    """Print a section title with formatting"""
    print(f"\n{Colors.BOLD}{Colors.BLUE}{'='*40}")
    print(f"{title}")
    print(f"{'='*40}{Colors.ENDC}")

def clean_text_for_excel(text):
    """Clean text to make it Excel-safe"""
    if not isinstance(text, str):
        return text
    # Remove control characters and special characters
    text = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', text)
    # Replace other problematic characters with space
    text = re.sub(r'[^\x00-\x7F]+', ' ', text)
    return text

def generate_extraction_report(pdf_directory):
    """
    Process all PDFs and generate both console report and Excel output with enhanced logging
    """
    print_section(f"Testing PDFs in directory: {pdf_directory}")
    
    # Get all PDF files in the directory
    pdf_files = [f for f in os.listdir(pdf_directory) if f.lower().endswith('.pdf')]
    
    if not pdf_files:
        print_colored("No PDF files found in the directory!", Colors.RED)
        return
        
    print_colored(f"Found {len(pdf_files)} PDF files", Colors.BLUE)
    print("-" * 80)
    
    # Initialize counters and data collection
    total_files = len(pdf_files)
    successful_extractions = 0
    partial_extractions = 0
    failed_extractions = 0
    extraction_results = []
    
    # Process each PDF
    for pdf_file in pdf_files:
        pdf_path = os.path.join(pdf_directory, pdf_file)
        print_section(f"Processing: {pdf_file}")
        
        result = {
            'PDF_File': pdf_file,
            'Account_Number': None,
            'Customer_Name': None,
            'Extraction_Status': 'Failed',
            'Error_Message': None,
            'Raw_Content': None,
            'Cleaned_Content': None,
            'Account_Line': None,
            'Name_Line': None
        }
        
        try:
            # Stage 1: Raw Extraction
            print("\n1. Raw Text Extraction:")
            print("-" * 20)
            raw_content = extract_header_content(pdf_path)
            result['Raw_Content'] = clean_text_for_excel(raw_content)
            print_colored(raw_content or "No raw content extracted", Colors.BLUE)
            
            # Stage 2: Content Cleanup
            print("\n2. Cleaned Content:")
            print("-" * 20)
            filtered_data = filter_header_data(raw_content)
            result['Cleaned_Content'] = clean_text_for_excel(filtered_data['cleaned_content'])
            print_colored(filtered_data['cleaned_content'] or "No cleaned content", Colors.YELLOW)
            
            # Stage 3: Final Extraction
            print("\n3. Final Extraction:")
            print("-" * 20)
            result['Account_Number'] = filtered_data['account_number']
            result['Customer_Name'] = filtered_data['customer_name']
            result['Account_Line'] = clean_text_for_excel(filtered_data['account_line'])
            result['Name_Line'] = clean_text_for_excel(filtered_data['name_line'])
            
            print("Account Number Line:", filtered_data['account_line'])
            print("Name Line:", filtered_data['name_line'])
            
            # Determine success based on finding either account number or name
            if filtered_data['account_number'] and filtered_data['customer_name']:
                result['Extraction_Status'] = 'Success'
                successful_extractions += 1
                status_color = Colors.GREEN
            elif filtered_data['account_number'] or filtered_data['customer_name']:
                result['Extraction_Status'] = 'Partial'
                partial_extractions += 1
                status_color = Colors.YELLOW
            else:
                result['Extraction_Status'] = 'Failed'
                result['Error_Message'] = 'No account number or customer name found'
                failed_extractions += 1
                status_color = Colors.RED
                
        except Exception as e:
            result['Error_Message'] = str(e)
            failed_extractions += 1
            status_color = Colors.RED
            
        extraction_results.append(result)
        
        # Print individual result with color
        print("\n4. Final Results:")
        print("-" * 20)
        print_colored(f"Status: {result['Extraction_Status']}", status_color)
        print(f"Account Number: {result['Account_Number']}")
        print(f"Customer Name: {result['Customer_Name']}")
        if result['Error_Message']:
            print_colored(f"Error: {result['Error_Message']}", Colors.RED)
        
        print("\n" + "-" * 80)
    
    # Print summary report
    print_section("EXTRACTION SUMMARY REPORT")
    print_colored(f"Total PDFs Processed: {total_files}", Colors.BLUE)
    print_colored(f"Successful Extractions: {successful_extractions}", Colors.GREEN)
    print_colored(f"Partial Extractions: {partial_extractions}", Colors.YELLOW)
    print_colored(f"Failed Extractions: {failed_extractions}", Colors.RED)
    print_colored(f"Success Rate: {((successful_extractions + partial_extractions)/total_files)*100:.2f}%", Colors.BLUE)
    
    # Create DataFrame and save to Excel
    df = pd.DataFrame(extraction_results)
    
    # Reorder columns for better readability
    columns_order = [
        'PDF_File', 
        'Extraction_Status', 
        'Account_Number', 
        'Account_Line',
        'Customer_Name', 
        'Name_Line',
        'Error_Message',
        'Cleaned_Content',
        'Raw_Content'
    ]
    df = df[columns_order]
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = os.path.join(pdf_directory, f'extraction_results_{timestamp}.xlsx')
    
    # Add color coding to Excel
    writer = pd.ExcelWriter(excel_file, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Results')
    
    # Get the workbook and the worksheet
    workbook = writer.book
    worksheet = writer.sheets['Results']
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap width at 50 characters
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Define styles
    success_fill = openpyxl.styles.PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    partial_fill = openpyxl.styles.PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    failed_fill = openpyxl.styles.PatternFill(start_color='FFB6C6', end_color='FFB6C6', fill_type='solid')
    
    # Apply conditional formatting
    status_col = df.columns.get_loc('Extraction_Status') + 1
    for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=status_col, max_col=status_col)):
        cell = row[0]
        if cell.value == 'Success':
            cell.fill = success_fill
        elif cell.value == 'Partial':
            cell.fill = partial_fill
        else:
            cell.fill = failed_fill
    
    writer.close()
    
    print(f"\nDetailed results saved to: {excel_file}")
    return df

if __name__ == "__main__":
    pdf_dir = r"C:\Users\Abcom\Desktop\extraction_name\pdfs"
    results_df = generate_extraction_report(pdf_dir)
